import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path

# Try image support via openpyxl (requires Pillow). Fallback gracefully if unavailable.
try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

BASE_DIR = Path(__file__).resolve().parent
DATA_CSV = BASE_DIR / "StudentsPerformance.csv"
OUT_XLSX = BASE_DIR / "Students_Performance_Dashboard.xlsx"
PLOTS_DIR = BASE_DIR / "plots"
PLOTS_DIR.mkdir(exist_ok=True)

# 1) Load data
usecols = None  # read all columns

df = pd.read_csv(DATA_CSV)
# Normalize column names
col_map = {
    'gender': 'Gender',
    'race/ethnicity': 'Race_Ethnicity',
    'parental level of education': 'Parental_Education',
    'lunch': 'Lunch',
    'test preparation course': 'Test_Prep',
    'math score': 'Math',
    'reading score': 'Reading',
    'writing score': 'Writing',
}
df = df.rename(columns=col_map)

# 2) Basic data quality checks
missing_counts = df.isna().sum().sort_values(ascending=False)
duplicate_count = int(df.duplicated().sum())

# 3) Fill nulls (if any) for numeric scores with column mean; for categoricals with mode
num_cols = ['Math', 'Reading', 'Writing']
cat_cols = [c for c in df.columns if c not in num_cols]

for c in num_cols:
    if df[c].isna().any():
        df[c] = df[c].fillna(df[c].mean())

for c in cat_cols:
    if df[c].isna().any():
        df[c] = df[c].fillna(df[c].mode().iloc[0])

# 4) Derived fields
# Average score and performance band similar to a classification
df['Average'] = df[num_cols].mean(axis=1).round(2)

bins = [ -np.inf, 50, 70, 85, np.inf ]
labels = ['Low', 'Fair', 'Good', 'Excellent']
df['Performance_Band'] = pd.cut(df['Average'], bins=bins, labels=labels)

# 5) Filtering + Sorting examples (keep format parity with BMW workflow tasks)
# - Keep full dataset; but prepare some commonly used filtered/sorted views
sorted_by_avg_desc = df.sort_values('Average', ascending=False).head(20)
sorted_by_math_desc = df.sort_values('Math', ascending=False).head(20)

# 6) Aggregations (analogous to sales-by-region etc.)
by_gender = df.groupby('Gender')[num_cols + ['Average']].mean().round(2)
by_race = df.groupby('Race_Ethnicity')[num_cols + ['Average']].mean().round(2)
by_prep = df.groupby('Test_Prep')[num_cols + ['Average']].mean().round(2)
band_counts = df['Performance_Band'].value_counts().sort_index()

# 7) Plots (Matplotlib)
plt.style.use('seaborn-v0_8-whitegrid')

# a) Average by Gender
ax = by_gender['Average'].plot(kind='bar', figsize=(6,4), color="#4F81BD", title='Average Score by Gender')
ax.set_ylabel('Average')
plt.tight_layout()
plot_gender = PLOTS_DIR / 'avg_by_gender.png'
plt.savefig(plot_gender, dpi=200)
plt.close()

# b) Average by Race/Ethnicity
ax = by_race['Average'].plot(kind='bar', figsize=(8,4), color="#5A5A5A", title='Average Score by Race/Ethnicity')
ax.set_ylabel('Average')
plt.tight_layout()
plot_race = PLOTS_DIR / 'avg_by_race.png'
plt.savefig(plot_race, dpi=200)
plt.close()

# c) Test Prep impact (completed vs none) per subject
ax = by_prep[num_cols + ['Average']].plot(kind='bar', figsize=(8,4), title='Scores by Test Prep Course')
ax.set_ylabel('Score')
plt.tight_layout()
plot_prep = PLOTS_DIR / 'scores_by_prep.png'
plt.savefig(plot_prep, dpi=200)
plt.close()

# d) Performance Band counts
ax = band_counts.plot(kind='bar', figsize=(6,4), color="#003399", title='Performance Band Distribution')
ax.set_ylabel('Count')
plt.tight_layout()
plot_band = PLOTS_DIR / 'band_counts.png'
plt.savefig(plot_band, dpi=200)
plt.close()

# 8) Export to Excel with two sheets (Cleaned Data + Summary)
# Use openpyxl so we can optionally embed images similar to the BMW workflow.
if OPENPYXL_OK:
    wb = Workbook()

    # Sheet 1: Cleaned Data
    ws1 = wb.active
    ws1.title = 'Cleaned_Data'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)

    # Sheet 2: Summary (quality checks, aggregations, top lists)
    ws2 = wb.create_sheet('Summary')

    def write_title(ws, title):
        ws.append([title])
        ws.append([])

    # Data quality
    write_title(ws2, 'Data Quality')
    ws2.append(['Duplicate Rows', duplicate_count])
    ws2.append([])
    ws2.append(['Missing Values (per column)'])
    ws2.append(['Column', 'Missing'])
    for c, v in missing_counts.items():
        ws2.append([c, int(v)])
    ws2.append([])

    # Aggregations
    write_title(ws2, 'Average Scores by Gender')
    ws2.append(['Gender'] + list(by_gender.columns))
    for idx, row in by_gender.iterrows():
        ws2.append([idx] + list(row.values))
    ws2.append([])

    write_title(ws2, 'Average Scores by Race/Ethnicity')
    ws2.append(['Race_Ethnicity'] + list(by_race.columns))
    for idx, row in by_race.iterrows():
        ws2.append([idx] + list(row.values))
    ws2.append([])

    write_title(ws2, 'Scores by Test Preparation')
    ws2.append(['Test_Prep'] + list(by_prep.columns))
    for idx, row in by_prep.iterrows():
        ws2.append([idx] + list(row.values))
    ws2.append([])

    write_title(ws2, 'Top 20 by Average')
    ws2.append(list(sorted_by_avg_desc.columns))
    for _, row in sorted_by_avg_desc.iterrows():
        ws2.append(list(row.values))
    ws2.append([])

    write_title(ws2, 'Top 20 by Math')
    ws2.append(list(sorted_by_math_desc.columns))
    for _, row in sorted_by_math_desc.iterrows():
        ws2.append(list(row.values))
    ws2.append([])

    # Embed plots if possible
    try:
        row_cursor = ws2.max_row + 2
        plots = [
            ('Average by Gender', plot_gender),
            ('Average by Race/Ethnicity', plot_race),
            ('Scores by Test Prep', plot_prep),
            ('Performance Bands', plot_band),
        ]
        for title, path in plots:
            ws2.cell(row=row_cursor, column=1, value=title)
            row_cursor += 1
            img = XLImage(str(path))
            img.width = img.width  # keep native
            img.height = img.height
            ws2.add_image(img, f"A{row_cursor}")
            row_cursor += 20  # spacing between charts
    except Exception:
        # If Pillow is not available or image add fails, just continue
        pass

    wb.save(OUT_XLSX)
else:
    # Fallback: use pandas ExcelWriter (without embedded images)
    with pd.ExcelWriter(OUT_XLSX, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cleaned_Data', index=False)

        # Build a summary workbook via pandas
        summary_blocks = []
        summary_blocks.append(pd.DataFrame({'Metric': ['Duplicate Rows'], 'Value': [duplicate_count]}))
        summary_blocks.append(missing_counts.rename('Missing').reset_index().rename(columns={'index': 'Column'}))
        summary_blocks.append(pd.DataFrame({'Section': ['Average by Gender']}))
        summary_blocks.append(by_gender.reset_index())
        summary_blocks.append(pd.DataFrame({'Section': ['Average by Race/Ethnicity']}))
        summary_blocks.append(by_race.reset_index())
        summary_blocks.append(pd.DataFrame({'Section': ['Scores by Test Prep']}))
        summary_blocks.append(by_prep.reset_index())
        summary_blocks.append(pd.DataFrame({'Section': ['Top 20 by Average']}))
        summary_blocks.append(sorted_by_avg_desc.reset_index(drop=True))
        summary_blocks.append(pd.DataFrame({'Section': ['Top 20 by Math']}))
        summary_blocks.append(sorted_by_math_desc.reset_index(drop=True))

        # Concatenate with blank separators
        out_df = []
        for block in summary_blocks:
            out_df.append(block)
            out_df.append(pd.DataFrame({'': []}))
        pd.concat(out_df, ignore_index=True).to_excel(writer, sheet_name='Summary', index=False)

print(f"Wrote: {OUT_XLSX}")
